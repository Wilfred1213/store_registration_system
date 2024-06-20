import datetime
from tempfile import NamedTemporaryFile

from django.db.models import Prefetch
from django.http import HttpResponse
from openpyxl import Workbook

from personas.models import Estudiante
from .models import Seccion


def exportar_datos_de_secciones(modeladmin, request, queryset):
    """
    Acción de Django Admin que exporta a Excel las secciones con la cantidad
    de estudiantes, niños y niñas de cada sección haciendo uso de la librería
    openpyxl ver: openpyxl.readthedocs.io
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Estadísticas de secciones"
    ws.append(["Sección", "Femenino", "Masculino", "Total"])
    ws.column_dimensions["A"].width = 50.30
    ws.column_dimensions["B"].width = 9.3
    ws.column_dimensions["C"].width = 9.3
    ws.column_dimensions["D"].width = 9.3
    count = 1
    for obj in queryset:
        ws.append(
            [
                str(obj).title(),
                obj.total_femenino(),
                obj.total_masculino(),
                obj.total_estudiantes(),
            ]
        )
        count += 1

    ws.append(
        [
            "Totales",
            "=SUM(B2:B" + str(count) + ")",
            "=SUM(C2:C" + str(count) + ")",
            "=SUM(D2:D" + str(count) + ")",
        ]
    )

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

        response = HttpResponse(
            content=stream,
            content_type="application/ms-excel",
        )
        response[
            "Content-Disposition"
        ] = f'attachment; filename=Cantidad_de_estudiantes_por_secciones-{datetime.datetime.now().strftime("%Y_%m_%d_%H-%M")}.xlsx'
        return response


def exportar_datos_de_grados(modeladmin, request, queryset):
    """
    Acción de Django Admin que exporta a Excel los Niveles Educativos con la
    cantidad de estudiantes, niños y niñas de cada sección haciendo uso de la
    librería openpyxl ver: openpyxl.readthedocs.io
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Estadísticas de grados"
    ws.append(["Grado", "Femenino", "Masculino", "Total"])
    ws.column_dimensions["A"].width = 50.30
    ws.column_dimensions["B"].width = 9.3
    ws.column_dimensions["C"].width = 9.3
    ws.column_dimensions["D"].width = 9.3
    count = 1
    for obj in queryset:
        ws.append(
            [
                str(obj).title(),
                obj.total_femenino(),
                obj.total_masculino(),
                obj.total_estudiantes(),
            ]
        )
        count += 1

    ws.append(
        [
            "Totales",
            "=SUM(B2:B" + str(count) + ")",
            "=SUM(C2:C" + str(count) + ")",
            "=SUM(D2:D" + str(count) + ")",
        ]
    )

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

        response = HttpResponse(
            content=stream,
            content_type="application/ms-excel",
        )
        response[
            "Content-Disposition"
        ] = f'attachment; filename=Cantidad_de_estudiantes_por_grados-{datetime.datetime.now().strftime("%Y_%m_%d_%H-%M")}.xlsx'
        return response


def exportar_a_excel_lista_de_firma_por_seccion(self, request, queryset):
    """
    Acción que toma un queryset de Secciones y exporta
    a un libro de excel una hoja de cálculo por sección con el nombre completo
    del estudiante, el nombre completo de su responsable, DUI y un espacio para firmas
    """
    secciones = Seccion.objects.filter(periodo_escolar__periodo_activo=True).prefetch_related(
        Prefetch(
            "estudiantes_en",
            queryset=Estudiante.objects.select_related("responsable").order_by(
                "apellidos", "nombre"
            ),
            to_attr="estudiantes",
        )
    ).order_by("nivel_educativo__edad_normal_de_ingreso")

    wb = Workbook()

    for seccion in secciones:
        ws = wb.create_sheet(seccion.__str__())
        ws.title = seccion.__str__()
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 50
        ws.column_dimensions["E"].width = 50
        ws.column_dimensions["F"].width = 50
        ws["A1"] = 'Complejo Educativo "Dr. Humberto Romero ALvergue"'
        ws.merge_cells("A1:F1")
        ws["A2"] = "Control de asistencia para padres de familia"
        ws.merge_cells("A2:F2")
        ws["A3"] = "Descripción de actividad"
        ws.merge_cells("A3:F3")
        ws["A4"] = ""
        ws.merge_cells("A5:F5")
        ws.append(["Nº", "Nombre de Padre /Madre o Responsable", "Nº de DUI", "Estudiante", "Sección de estudiante", "Firma"])
        counter = 1
        for estudiante in seccion.estudiantes_en.all():
            ws.append([str(counter), estudiante.responsable.__str__(),estudiante.__str__(), seccion.__str__(), ""])
            counter += 1
    


    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

        response = HttpResponse(
            content=stream,
            content_type="application/ms-excel",
        )
        response[
            "Content-Disposition"
        ] = f'attachment; filename=ListaDeEstudiantesYResponsables-{datetime.datetime.now().strftime("%Y_%m_%d-%H%M")}.xlsx'
        return response
